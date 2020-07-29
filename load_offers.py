#!/usr/bin/env python3

import re
from weakref import WeakKeyDictionary

import bs4
import requests
import openpyxl
import openpyxl.utils as xlutil

INTERNET_URL="http://kungsbacka.openuniverse.se/kungsbackastadsnat/internet/privat/"
INTERNET_FILE="internet.html"
PAKET_URL="http://kungsbacka.openuniverse.se/kungsbackastadsnat/paket/privat/"
PAKET_FILE="paket.html"

XLS_FILE="priser.xlsx"


def get_webpages():
    r = requests.get(INTERNET_URL)
    assert (r.status_code == requests.codes.ok)
    with open(INTERNET_FILE, "w") as fp:
        fp.write(r.text)


class IntProp:
    def __init__(self):
        self.values = WeakKeyDictionary()

    def __get__(self, instance, owner):
        return self.values.get(instance, None)

    def __set__(self, instance, value):
        self.values[instance] = int(value)


class Offer:
    def __init__(self):
        self.name = None
        self.banner = None
        self.isp = None
        self.campaign = None
        self.bind_time = None
        self.leave_time = None

    list_price = IntProp()
    speed_up = IntProp()
    speed_down = IntProp()
    start_cost = IntProp()

    @property
    def year_cost(self):
        if not self.campaign:
            return self.start_cost + self.list_price*12

        months = re.search(r"\b(\d\d?|tre) ?mån", self.campaign, flags=re.IGNORECASE)
        if months:  # halva priset eller xx kr/månad i antal månader
            months = months.group(1)
            if months == "tre":
                months = 3
            try:
                months = int(months)
            except TypeError:
                print(self.campaign)
                raise
        elif "ett halvår" in self.campaign:
            months = 6
        else:  # Vi lyckas inte lista ut den här kampanjen
            print(str(self) + " Unhandled campaign: " + self.campaign)
            self.campaign = "## Campaign parse failed ## " + self.campaign

            return self.start_cost + self.list_price*12
        reduced = re.search(r"(\d+) ?kr", self.campaign, flags=re.IGNORECASE)
        if reduced:  # justerat pris i kronor
            reduced_price = int(reduced.group(1))
            if reduced_price:
                return months*reduced_price + self.list_price*(12-months) + self.start_cost
        return self.list_price*(months*0.5 + 12 - months) + self.start_cost

    def __str__(self):
        r = self.isp
        r += "\n\tSpeed: " + str(self.speed_up) + "/" + str(self.speed_down)
        return r


def parse_page(file):
    with open(file, "r") as fp:
        soup = bs4.BeautifulSoup(fp.read(), features="html.parser")
    offers = []
    speed_re = re.compile(r"(\d+)[_/](\d+)")
    for offer in soup.select("article.serviceOffer"):
        o = Offer()
        o.banner = offer.select(".serviceOfferText > h4")[0].getText().strip()
        o.name = offer.select(".product-name")[0].getText()
        o.isp = offer.select(".isp-link")[0].attrs["href"].split("/")[3]

        speed = speed_re.search(o.name) or speed_re.search(o.banner)
        if speed:
            o.speed_down = speed.group(1)
            o.speed_up = speed.group(2)
        else:
            speed = re.search(r"\d+", o.banner)
            o.speed_up = speed.group(0)
            o.speed_down = speed.group(0)
        o.list_price = offer.select("strong[data-bdd-id='endUserPrice']")[0].get_text().split()[0]
        if "campaign" in offer["class"]:
            c = offer.select("div[data-bdd-id='campaignDetailsText']")[0]
            o.campaign = c.get_text().strip().split("\n", maxsplit=1)[1].strip()

        conditions = offer.select("p[data-bdd-id='serviceOfferPurchaseInformation']")[0]
        (s, b, l) = conditions.get_text().strip().split("\n")
        o.start_cost = s.split()[1].strip()
        o.bind_time = b.split(":")[1].strip()
        o.leave_time = l.split(":")[1].strip()
        offers.append(o)

    return offers


def write_xslx(offers):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Internet"
    # Title, Offer property, column width, optional number format
    headers = [("ISP",          "isp",              18),
               ("Downlink",     "speed_down",       18),
               ("Uplink",       "speed_up",         18),
               ("Månadskostnad", "list_price",      20, "# ##0,00 [$kr-41D]"),
               ("Kampanj",      "campaign",         4.6*12),
               ("Startkostnad", "start_cost",       16, "# ##0 [$kr-41D]"),
               ("Bindning",     "bind_time",        16),
               # ("Uppsägning", "leave_time",     16),
               ("Årskostnad",   "year_cost",        16, "# ##0 [$kr-41D]"),
               ]
    for col, hdr in enumerate(headers, start=1):
        sheet.cell(1, col, hdr[0])
        sheet.column_dimensions[xlutil.cell.get_column_letter(col)].width = hdr[2]
    for row, offer in enumerate(offers, start=2):
        for col, hdr in enumerate(headers, start=1):
            value = getattr(offer, hdr[1])
            sheet.cell(row, col, value)
            if len(hdr) > 3:
                sheet.cell(row, col).number_format = hdr[3]

    sheet.auto_filter.ref = "A1:" + xlutil.cell.get_column_letter(len(headers)) + str(sheet.max_row)
    wb.save(XLS_FILE)


def main():
    get_webpages()
    offers = parse_page(INTERNET_FILE)
    write_xslx(offers)


if __name__ == "__main__":
    main()
